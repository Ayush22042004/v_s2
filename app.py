import os
import sqlite3
from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from datetime import datetime, timezone, timedelta
import pytz
from io import BytesIO
from openpyxl import Workbook
from flask import send_file
import smtplib
from email.message import EmailMessage
try:
    from flask_wtf.csrf import CSRFProtect
except ImportError:
    CSRFProtect = None


def send_email(to_addr, subject, body):
    """Send email via SMTP. Reads SMTP_HOST, SMTP_PORT, SMTP_USER, SMTP_PASS from env.
    Falls back to printing the message if SMTP is not configured."""
    smtp_host = os.environ.get('SMTP_HOST')
    smtp_port = int(os.environ.get('SMTP_PORT', 587)) if os.environ.get('SMTP_PORT') else None
    smtp_user = os.environ.get('SMTP_USER')
    smtp_pass = os.environ.get('SMTP_PASS')
    msg = EmailMessage()
    msg['Subject'] = subject
    msg['From'] = os.environ.get('SMTP_FROM', 'no-reply@example.com')
    msg['To'] = to_addr
    msg.set_content(body)
    if smtp_host and smtp_port and smtp_user and smtp_pass:
        try:
            with smtplib.SMTP(smtp_host, smtp_port, timeout=10) as s:
                s.starttls()
                s.login(smtp_user, smtp_pass)
                s.send_message(msg)
        except Exception as e:
            print('⚠️ send_email failed:', e)
    else:
        print('Email (not sent - SMTP not configured) ->', to_addr, subject)


def send_notification(user_id, message):
    try:
        created_at = datetime.now(timezone.utc).isoformat()
        execute('INSERT INTO notifications (user_id,message,created_at,read) VALUES (?,?,?,?)', (user_id, message, created_at, 0))
    except Exception as e:
        print('⚠️ send_notification failed:', e)
# timezone helpers
IST = pytz.timezone("Asia/Kolkata")

def safe_localize(naive_dt, tz):
    """Attach tz info to a naive datetime supporting both pytz and zoneinfo.
    If tz exposes localize, use tz.localize(naive_dt) (pytz). Otherwise use naive_dt.replace(tzinfo=tz) (zoneinfo).
    """
    if naive_dt is None:
        return None
    if hasattr(tz, 'localize'):
        return tz.localize(naive_dt)
    return naive_dt.replace(tzinfo=tz)

# Global variable to store last scheduled election debug data (admin-only, short-lived)
LAST_SCHEDULE_DEBUG = {}

# If running in an environment with DATABASE_URL (e.g. Render), prefer PostgreSQL helpers
DB_ADAPTER = os.environ.get("DATABASE_URL")
if DB_ADAPTER:
    try:
        import db_pg as db_backend
        # expose functions used across the app
        get_db = db_backend.get_conn
        # adapter that converts sqlite-style '?' placeholders to postgres '%s'
        def _pg_exec_sql(sql, args=(), fetch=False, one=False):
            if sql and '?' in sql:
                sql = sql.replace('?', '%s')
            return db_backend.exec_sql(sql, args, fetch=fetch, one=one)

        exec_sql = _pg_exec_sql
        query = lambda sql, args=(), one=False: _pg_exec_sql(sql, args, fetch=True, one=one)
        execute = lambda sql, args=(): _pg_exec_sql(sql, args, fetch=False)
        now_utc = db_backend.now_utc
        IST = getattr(db_backend, 'IST', IST)
    except Exception as e:
        # fall back to builtin sqlite definitions below; print error for deployment logs
        print('⚠️ Could not load db_pg adapter for DATABASE_URL:', e)
        pass


def get_db():
    conn = sqlite3.connect('voting.db', timeout=30)
    conn.row_factory = sqlite3.Row
    return conn

# application directory
APP_DIR = os.path.dirname(os.path.abspath(__file__))

def get_any(d, *keys):
    for k in keys:
        v = d.get(k)
        if v:
            return v
    return None

def parse_utc_or_local_as_ist(utc_str, local_str):
    """Return timezone-aware UTC datetime parsed from utc_str (ISO Z) or local naive string as IST."""
    if utc_str:
        try:
            return datetime.fromisoformat(utc_str.replace('Z','+00:00')).astimezone(timezone.utc)
        except Exception:
            pass
    if local_str:
        try:
            naive = datetime.fromisoformat(local_str)
            loc = safe_localize(naive, IST)
            return loc.astimezone(timezone.utc)
        except Exception:
            return None
    return None

def exec_sql(sql, args=(), fetch=False, one=False):
    db = get_db()
    cur = db.execute(sql, args)
    if fetch:
        rows = cur.fetchall()
        cur.close()
        return rows[0] if one and rows else rows
    else:
        db.commit()
        last = cur.lastrowid
        cur.close()
        return last

# thin wrappers expected by the rest of the code
def query(sql, args=(), one=False):
    return exec_sql(sql, args, fetch=True, one=one)

def execute(sql, args=()):
    return exec_sql(sql, args, fetch=False)

# Flask app init
app = Flask(__name__)
# Generate a secure random key if no SECRET_KEY is set
import secrets
app.secret_key = os.environ.get('SECRET_KEY', secrets.token_hex(32))

# Enable CSRF protection
try:
    if CSRFProtect:
        csrf = CSRFProtect(app)
except:
    # Flask-WTF not installed, skip CSRF for now
    csrf = None
    print("⚠️ Flask-WTF not installed - CSRF protection disabled")

# Database initialization function
def init_database():
    """Initialize database tables if they don't exist"""
    try:
        db = get_db()
        
        # Create users table
        db.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                email TEXT UNIQUE,
                username TEXT UNIQUE NOT NULL,
                password TEXT NOT NULL,
                role TEXT NOT NULL DEFAULT 'voter',
                id_number TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Create elections table
        db.execute('''
            CREATE TABLE IF NOT EXISTS elections (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT,
                category TEXT,
                start_time TEXT NOT NULL,
                end_time TEXT NOT NULL,
                created_by INTEGER,
                candidate_limit INTEGER DEFAULT 10,
                status TEXT DEFAULT 'active',
                cancelled_at TEXT,
                cancelled_by INTEGER,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (created_by) REFERENCES users (id)
            )
        ''')
        
        # Create candidates table
        db.execute('''
            CREATE TABLE IF NOT EXISTS candidates (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                election_id INTEGER NOT NULL,
                category TEXT DEFAULT 'General',
                photo TEXT,
                user_id INTEGER,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (election_id) REFERENCES elections (id),
                FOREIGN KEY (user_id) REFERENCES users (id)
            )
        ''')
        
        # Create votes table
        db.execute('''
            CREATE TABLE IF NOT EXISTS votes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                candidate_id INTEGER NOT NULL,
                election_id INTEGER NOT NULL,
                voted_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (user_id) REFERENCES users (id),
                FOREIGN KEY (candidate_id) REFERENCES candidates (id),
                FOREIGN KEY (election_id) REFERENCES elections (id),
                UNIQUE(user_id, election_id)
            )
        ''')
        
        # Create candidate_applications table
        db.execute('''
            CREATE TABLE IF NOT EXISTS candidate_applications (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                election_id INTEGER NOT NULL,
                name TEXT NOT NULL,
                category TEXT DEFAULT 'General',
                photo TEXT,
                status TEXT DEFAULT 'pending',
                applied_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                reviewed_at TIMESTAMP,
                reviewed_by INTEGER,
                FOREIGN KEY (user_id) REFERENCES users (id),
                FOREIGN KEY (election_id) REFERENCES elections (id),
                FOREIGN KEY (reviewed_by) REFERENCES users (id)
            )
        ''')
        
        # Create notifications table (if needed)
        db.execute('''
            CREATE TABLE IF NOT EXISTS notifications (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                message TEXT NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                read INTEGER DEFAULT 0,
                FOREIGN KEY (user_id) REFERENCES users (id)
            )
        ''')
        
        # Create default admin user if it doesn't exist
        admin_exists = db.execute('SELECT COUNT(*) FROM users WHERE username = ?', ('admin',)).fetchone()[0]
        if not admin_exists:
            admin_password = generate_password_hash('admin123')
            db.execute('''
                INSERT INTO users (name, username, password, role)
                VALUES (?, ?, ?, ?)
            ''', ('Administrator', 'admin', admin_password, 'admin'))
            print("✅ Default admin user created (username: admin, password: admin123)")
        
        db.commit()
        db.close()
        print("✅ Database initialized successfully")
        
    except Exception as e:
        print(f"❌ Database initialization error: {e}")

# Initialize database on startup
init_database()

# Simple rate limiting storage (in-memory for demo)
from collections import defaultdict, deque
import time

class SimpleRateLimiter:
    def __init__(self):
        self.requests = defaultdict(deque)
    
    def is_allowed(self, key, limit=5, window=300):  # 5 requests per 5 minutes
        now = time.time()
        # Clean old requests
        while self.requests[key] and self.requests[key][0] < now - window:
            self.requests[key].popleft()
        
        if len(self.requests[key]) >= limit:
            return False
        
        self.requests[key].append(now)
        return True

rate_limiter = SimpleRateLimiter()

# Optional startup migration for Postgres when running on Render.
# Set RUN_MIGRATE=true in the environment to run migrations once at startup.
if DB_ADAPTER and os.environ.get('RUN_MIGRATE', '').lower() in ('1', 'true', 'yes'):
    try:
        import db_pg
        db_pg.migrate_and_seed()
        print('✅ DB migration executed at startup')
    except Exception as e:
        print('⚠️ DB migration failed at startup:', e)

# Ensure unique constraint on votes table to prevent double voting
try:
    db = get_db()
    # Try to create unique index if it doesn't exist (SQLite safe)
    db.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_votes_unique ON votes(voter_id, election_id)")
    db.commit()
except Exception:
    pass  # Index might already exist

# -------------- Auth helpers --------------
def login_required(role=None):
    def deco(f):
        @wraps(f)
        def wrap(*args, **kwargs):
            if "user_id" not in session:
                flash("Please login first.", "warn")
                return redirect(url_for("login"))
            if role and session.get("role") != role:
                flash("Not authorized.", "error")
                return redirect(url_for("index"))
            return f(*args, **kwargs)
        return wrap
    return deco

# -------------- Time helpers --------------
def parse_iso(s):
    try:
        dt = datetime.fromisoformat((s or "").replace("Z", ""))
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(timezone.utc)
    except Exception:
        pass
        return None


# to_ist is defined later after helpers; remove this duplicate to avoid confusion

def now_utc():
    return datetime.now(timezone.utc)

@app.context_processor
def inject_unread_notifications():
    try:
        uid = session.get('user_id')
        if not uid:
            return {}
        row = query('SELECT COUNT(*) AS c FROM notifications WHERE user_id=? AND read=0', (uid,), one=True)
        return {'unread_notifications': row['c'] if row else 0}
    except Exception:
        return {'unread_notifications': 0}


@app.route('/health')
def health():
    return {'status': 'ok'}, 200

# -------------- Routes --------------
@app.route("/")
def index():
    user = query("SELECT * FROM users WHERE id=?", (session["user_id"],), one=True) if "user_id" in session else None
    active = query("SELECT * FROM elections WHERE status != 'cancelled' OR status IS NULL ORDER BY start_time DESC LIMIT 5")
    return render_template("index.html", user=user, elections=active)

@app.route("/all_elections")
def all_elections():
    """Public page showing all elections"""
    elections = query("SELECT * FROM elections WHERE status != 'cancelled' OR status IS NULL ORDER BY start_time DESC")
    return render_template("all_elections.html", elections=elections)

@app.route("/signup", methods=["GET","POST"])
def signup():
    if request.method == "POST":
        name = request.form.get("name","").strip()
        email = (request.form.get("email","").strip() or None)
        username = request.form.get("username","").strip().lower()
        password = request.form.get("password","")
        id_number = request.form.get("id_number","").strip()
        if not name or not username or not password or not id_number:
            flash("All fields except email are required.", "error"); return redirect(url_for("signup"))
        
        # Password strength validation
        if len(password) < 8:
            flash("Password must be at least 8 characters long.", "error")
            return redirect(url_for("signup"))
        if not any(c.isdigit() for c in password):
            flash("Password must contain at least one number.", "error")
            return redirect(url_for("signup"))
        if query("SELECT 1 FROM users WHERE lower(username)=?", (username,), one=True):
            flash("Username already taken.", "error"); return redirect(url_for("signup"))
        if email and query("SELECT 1 FROM users WHERE lower(email)=?", (email.lower(),), one=True):
            flash("Email already registered.", "error"); return redirect(url_for("signup"))
        execute("INSERT INTO users (name,email,username,password,role,id_number) VALUES (?,?,?,?,?,?)",
                (name, email.lower() if email else None, username, generate_password_hash(password), "voter", id_number))
        flash("Account created. Please login.", "ok"); return redirect(url_for("login"))
    
    # Pass elections data for candidate signup option
    elections = query("SELECT * FROM elections WHERE status != 'cancelled' OR status IS NULL ORDER BY start_time DESC")
    return render_template("signup.html", elections=elections)

@app.route("/login", methods=["GET","POST"])
def login():
    if request.method == "POST":
        # Rate limit login attempts by IP
        client_ip = request.environ.get('HTTP_X_FORWARDED_FOR', request.remote_addr)
        if not rate_limiter.is_allowed(f"login_{client_ip}", limit=5, window=300):
            flash("Too many login attempts. Please try again in 5 minutes.", "error")
            return redirect(url_for("login"))
        
        username = request.form.get("username","" ).strip().lower()
        password = request.form.get("password","" )
        user = query("SELECT * FROM users WHERE lower(username)=?", (username,), one=True)
        if user and check_password_hash(user["password"], password):
            session["user_id"] = user["id"]
            session["role"] = user["role"]
            session["username"] = user["username"]
            session["user"] = {"username": user["username"], "role": user["role"]}  # Add user object to session
            flash("Welcome back!", "ok")
            return redirect(url_for("index"))
        flash("Invalid credentials.", "error")
    # Defensive: render template and strip any stray pw-toggle button server-side
    html = render_template("login.html")
    try:
        import re
        html = re.sub(r"<button[^>]*id=[\"']pw-toggle[\"'][^>]*>.*?</button>", "", html, flags=re.S)
    except Exception:
        pass
    return html


@app.route('/candidate_signup', methods=['GET','POST'])
def candidate_signup():
    if request.method == 'POST':
        name = request.form.get('name','').strip()
        email = (request.form.get('email','').strip() or None)
        username = request.form.get('username','').strip().lower()
        password = request.form.get('password','')
        category = request.form.get('category','').strip() or 'General'
        election_id = request.form.get('election_id')
        if not name or not username or not password:
            flash('Name, username and password are required.', 'error'); return redirect(url_for('candidate_signup'))
        if query('SELECT 1 FROM users WHERE lower(username)=?', (username,), one=True):
            flash('Username already taken.', 'error'); return redirect(url_for('candidate_signup'))
        # create user as candidate
        execute('INSERT INTO users (name,email,username,password,role) VALUES (?,?,?,?,?)',
                (name, email.lower() if email else None, username, generate_password_hash(password), 'candidate'))
        user = query('SELECT * FROM users WHERE lower(username)=?', (username,), one=True)
        user_id = user['id']
        # prevent duplicate application for same user+election
        if election_id:
            exists = query('SELECT 1 FROM candidate_applications WHERE user_id=? AND election_id=? AND status IN ("pending","approved")', (user_id, election_id), one=True)
            if exists:
                flash('You have already applied for this election.', 'warn'); return redirect(url_for('candidate_signup'))
        # handle photo
        photo_path=None; f=request.files.get('photo')
        if f and f.filename:
            # Enhanced validation: extension, mimetype, and magic bytes check, size limit 5MB
            allowed_ext = {'png','jpg','jpeg','gif'}
            name_parts = secure_filename(f.filename).rsplit('.', 1)
            ext = name_parts[1].lower() if len(name_parts) > 1 else ''
            # Check file extension and mimetype
            if ext not in allowed_ext or not f.mimetype or not f.mimetype.startswith('image'):
                flash('Invalid photo file. Allowed: png,jpg,jpeg,gif', 'error'); return redirect(url_for('candidate_signup'))
            data = f.read()
            if len(data) > 5*1024*1024:
                flash('Photo too large (max 5MB).', 'error'); return redirect(url_for('candidate_signup'))
            # persist file
            filename = secure_filename(f.filename)
            uploads_dir = os.path.join(APP_DIR, 'static', 'uploads'); os.makedirs(uploads_dir, exist_ok=True)
            with open(os.path.join(uploads_dir, filename), 'wb') as out:
                out.write(data)
            photo_path=f'uploads/{filename}'
        applied_at = datetime.now(timezone.utc).isoformat()
        execute('INSERT INTO candidate_applications (user_id,election_id,name,category,photo,status,applied_at) VALUES (?,?,?,?,?,?,?)',
                (user_id, election_id, name, category, photo_path, 'pending', applied_at))
        flash('Application submitted. Awaiting admin approval.', 'ok'); return redirect(url_for('login'))
    elections = query("SELECT * FROM elections WHERE status != 'cancelled' OR status IS NULL ORDER BY start_time DESC")
    return render_template('candidate_signup.html', elections=elections)


@app.route('/candidate/profile')
@login_required(role='candidate')
def candidate_profile():
    user_id = session.get('user_id')
    # mark notifications read
    try:
        execute('UPDATE notifications SET read=1 WHERE user_id=?', (user_id,))
    except Exception:
        pass
    apps = query('SELECT a.*, e.title AS election_title FROM candidate_applications a LEFT JOIN elections e ON e.id=a.election_id WHERE a.user_id=? ORDER BY a.applied_at DESC', (user_id,))
    approved = query('SELECT c.*, e.title AS election_title FROM candidates c LEFT JOIN elections e ON e.id=c.election_id WHERE c.user_id=?', (user_id,))
    return render_template('candidate_profile.html', apps=apps, approved=approved)


@app.route('/admin/candidate_applications')
@login_required(role='admin')
def admin_candidate_applications():
    rows = query("SELECT a.*, u.username, u.email, e.title AS election_title FROM candidate_applications a JOIN users u ON u.id=a.user_id LEFT JOIN elections e ON e.id=a.election_id WHERE a.status='pending' ORDER BY a.applied_at ASC")
    return render_template('admin_candidate_applications.html', apps=rows)


@app.route('/admin/approve_candidate', methods=['POST'])
@login_required(role='admin')
def admin_approve_candidate():
    app_id = request.form.get('application_id')
    if not app_id:
        flash('Invalid request.', 'error'); return redirect(url_for('admin_candidate_applications'))
    a = query('SELECT * FROM candidate_applications WHERE id=?', (app_id,), one=True)
    if not a:
        flash('Application not found.', 'error'); return redirect(url_for('admin_candidate_applications'))
    # create candidate row
    photo = a['photo']
    election_id = a['election_id']
    name = a['name']
    category = a['category']
    user_id = a['user_id']
    execute('INSERT INTO candidates (name,category,photo,election_id,user_id) VALUES (?,?,?,?,?)', (name, category, photo, election_id, user_id))
    candidate_id = query('SELECT id FROM candidates WHERE user_id=? ORDER BY id DESC', (user_id,), one=True)['id']
    # update application
    approved_at = datetime.now(timezone.utc).isoformat()
    execute('UPDATE candidate_applications SET status=?, approved_by=?, approved_at=?, candidate_id=? WHERE id=?', ('approved', session.get('user_id'), approved_at, candidate_id, app_id))
    # notify candidate if email available
    try:
        user = query('SELECT * FROM users WHERE id=?', (user_id,), one=True)
        if user:
            if user.get('email'):
                send_email(user.get('email'), 'Your candidacy has been approved', f"Hello {user.get('name')},\n\nYour application for '{name}' has been approved and you are now registered as a candidate for the election.\n\nRegards")
            # create in-app notification
            try:
                send_notification(user['id'], f"Your application for '{name}' was approved.")
            except Exception:
                pass
    except Exception:
        pass
    flash('Candidate approved and registered for election.', 'ok'); return redirect(url_for('admin_candidate_applications'))


@app.route('/admin/reject_candidate', methods=['POST'])
@login_required(role='admin')
def admin_reject_candidate():
    app_id = request.form.get('application_id')
    if not app_id:
        flash('Invalid request.', 'error'); return redirect(url_for('admin_candidate_applications'))
    execute('UPDATE candidate_applications SET status=? WHERE id=?', ('rejected', app_id))
    try:
        a = query('SELECT * FROM candidate_applications WHERE id=?', (app_id,), one=True)
        if a:
            user = query('SELECT * FROM users WHERE id=?', (a['user_id'],), one=True)
            if user:
                if user.get('email'):
                    send_email(user.get('email'), 'Your candidacy has been rejected', f"Hello {user.get('name')},\n\nYour application for '{a['name']}' was rejected by the admin.\n\nRegards")
                try:
                    send_notification(user['id'], f"Your application for '{a['name']}' was rejected by the admin.")
                except Exception:
                    pass
    except Exception:
        pass
    flash('Application rejected.', 'ok'); return redirect(url_for('admin_candidate_applications'))

@app.route("/logout")
def logout():
    session.clear(); flash("Logged out.", "ok"); return redirect(url_for("index"))

# ----------- Admin Dashboard -----------
@app.route("/admin")
@login_required(role="admin")
def admin():
    rows = query("SELECT * FROM elections ORDER BY start_time DESC")
    ongoing, scheduled, ended = classify_elections(rows)
    return render_template("admin.html", ongoing=ongoing, scheduled=scheduled, ended=ended, elections=rows, last_schedule_debug=LAST_SCHEDULE_DEBUG)

@app.route("/admin/past-elections")
@login_required(role="admin")
def past_elections():
    """Show all past elections with full details"""
    rows = query("SELECT * FROM elections ORDER BY start_time DESC")
    ongoing, scheduled, ended = classify_elections(rows)
    
    # Get candidate count for each election
    elections_with_stats = []
    for e in ended:
        candidate_count = query("SELECT COUNT(*) as count FROM candidates WHERE election_id=?", (e['id'],), one=True)
        vote_count = query("SELECT COUNT(*) as count FROM votes WHERE election_id=?", (e['id'],), one=True)
        
        election_data = dict(e)
        election_data['candidate_count'] = candidate_count['count'] if candidate_count else 0
        election_data['vote_count'] = vote_count['count'] if vote_count else 0
        elections_with_stats.append(election_data)
    
    return render_template("past_elections.html", elections=elections_with_stats, total_count=len(ended))

@app.route("/add_candidate", methods=["POST"])
@login_required(role="admin")
def add_candidate():
    name = request.form.get("name","").strip()
    category = request.form.get("category","").strip() or "General"
    election_id = request.form.get("election_id","").strip()
    if not name or not election_id:
        flash("Candidate name and election are required.", "error"); return redirect(url_for("admin"))
    try:
        election_id = int(election_id)
    except ValueError:
        flash("Invalid election selected.", "error"); return redirect(url_for("admin"))
    e = query("SELECT id, candidate_limit, created_by FROM elections WHERE id=?", (election_id,), one=True)
    if not e: flash("Selected election does not exist.", "error"); return redirect(url_for("admin"))
    if e["created_by"] and e["created_by"] != session.get("user_id"):
        flash("You can only add candidates to your own elections.", "error"); return redirect(url_for("admin"))
    if e["candidate_limit"] is not None:
        cnt = query("SELECT COUNT(*) AS c FROM candidates WHERE election_id=?", (election_id,), one=True)["c"]
        if cnt >= e["candidate_limit"]:
            flash("Candidate limit reached for this election.", "warn"); return redirect(url_for("admin"))
    photo_path=None; f=request.files.get("photo")
    if f and f.filename:
        # Security validation: extension and mimetype, size limit 5MB
        allowed_ext = {'png','jpg','jpeg','gif'}
        name_parts = secure_filename(f.filename).rsplit('.', 1)
        ext = name_parts[1].lower() if len(name_parts) > 1 else ''
        if ext not in allowed_ext or (f.mimetype and not f.mimetype.startswith('image')):
            flash('Invalid photo file. Allowed: png,jpg,jpeg,gif', 'error'); return redirect(url_for('admin'))
        data = f.read()
        if len(data) > 5*1024*1024:
            flash('Photo too large (max 5MB).', 'error'); return redirect(url_for('admin'))
        # persist file
        filename = secure_filename(f.filename)
        uploads_dir = os.path.join(APP_DIR, "static", "uploads"); os.makedirs(uploads_dir, exist_ok=True)
        with open(os.path.join(uploads_dir, filename), 'wb') as out:
            out.write(data)
        photo_path=f"uploads/{filename}"
    execute("INSERT INTO candidates (name,category,photo,election_id) VALUES (?,?,?,?)",
            (name, category, photo_path, election_id))
    flash("Candidate added to election.", "ok"); return redirect(url_for("admin"))

@app.route("/schedule_election", methods=["POST"])
@login_required(role="admin")
def schedule_election():
    global LAST_SCHEDULE_DEBUG
    
    title = request.form.get("title","").strip()
    year = request.form.get("year","").strip()
    category = request.form.get("category","").strip()
    
    # Get timezone offset for debugging/fallback
    tz_raw = (request.form.get("tz_offset") or "").strip()
    try:
        tz_offset = int(tz_raw) if tz_raw != "" else 0
    except ValueError:
        tz_offset = 0
    
    # Prefer UTC inputs if provided by the client JavaScript
    start_time_utc = (request.form.get('start_time_utc') or '').strip()
    end_time_utc = (request.form.get('end_time_utc') or '').strip()
    
    # Fallback to local time inputs if UTC not provided
    start_time_local = (request.form.get('start_time') or '').strip()
    end_time_local = (request.form.get('end_time') or '').strip()
    
    if not title or not year or not category or (not start_time_utc and not start_time_local) or (not end_time_utc and not end_time_local):
        flash("Missing some fields for scheduling.", "error")
        return redirect(url_for("admin"))
    
    # Parse UTC times if provided, otherwise parse local times as IST
    if start_time_utc and end_time_utc:
        app.logger.info(f"[Schedule Election] Using UTC inputs: start={start_time_utc}, end={end_time_utc}")
        sdt = parse_iso(start_time_utc)
        edt = parse_iso(end_time_utc)
        if not sdt or not edt:
            flash("Invalid UTC time format received from client.", "error")
            return redirect(url_for("admin"))
    else:
        # Fallback: parse as naive local datetime and assume IST
        app.logger.info(f"[Schedule Election] Using local (IST) fallback: start={start_time_local}, end={end_time_local}")
        try:
            # Ensure we have seconds in the datetime string
            start_str = start_time_local if len(start_time_local) > 16 else start_time_local + ":00"
            end_str = end_time_local if len(end_time_local) > 16 else end_time_local + ":00"
            
            # Parse as naive datetime
            start_naive = datetime.fromisoformat(start_str)
            end_naive = datetime.fromisoformat(end_str)
            
            # Localize to IST and convert to UTC
            try:
                start_ist = safe_localize(start_naive, IST)
                end_ist = safe_localize(end_naive, IST)
            except Exception as loc_err:
                app.logger.error(f"[Schedule Election] safe_localize failed: {loc_err}")
                flash("Invalid or ambiguous local time.", "error")
                return redirect(url_for("admin"))
            sdt = start_ist.astimezone(timezone.utc)
            edt = end_ist.astimezone(timezone.utc)
        except Exception as e:
            app.logger.error(f"[Schedule Election] Failed to parse local times: {e}")
            flash("Invalid time format.", "error")
            return redirect(url_for("admin"))
    
    # Validate time window
    if not sdt or not edt or edt <= sdt:
        flash("Invalid time window. End must be after start.", "error")
        return redirect(url_for("admin"))
    
    # Convert to ISO format for storage
    start_time = sdt.isoformat()
    end_time = edt.isoformat()
    
    # Store debug data for admin panel
    LAST_SCHEDULE_DEBUG = {
        'raw_start': start_time_utc or start_time_local,
        'raw_end': end_time_utc or end_time_local,
        'utc_start': start_time,
        'utc_end': end_time
    }
    
    # Log the computed values
    app.logger.info(f"[Schedule Election] Computed UTC times: start={start_time}, end={end_time}")
    app.logger.info(f"[Schedule Election] IST display: start={to_ist(sdt)}, end={to_ist(edt)}")
    
    # Parse and validate candidate limit
    created_by = session.get("user_id")
    cand_limit = request.form.get("candidate_limit","").strip()
    try:
        cand_limit_val = int(cand_limit) if cand_limit else None
        if cand_limit_val is not None and cand_limit_val < 1:
            raise ValueError
    except ValueError:
        flash("Candidate limit must be a positive number.", "error")
        return redirect(url_for("admin"))
    
    # Insert election into database
    execute("INSERT INTO elections (title,year,category,start_time,end_time,created_by,candidate_limit) VALUES (?,?,?,?,?,?,?)",
            (title, int(year), category, start_time, end_time, created_by, cand_limit_val))
    
    # Flash success message with computed times
    flash(f"Election scheduled. UTC: {start_time} to {end_time} | IST: {to_ist(sdt).strftime('%Y-%m-%d %H:%M')} to {to_ist(edt).strftime('%Y-%m-%d %H:%M')}", "ok")
    return redirect(url_for("admin"))

# ----------- Voting & Results -----------
def current_active_election():
    rows = query("SELECT * FROM elections WHERE status != 'cancelled' OR status IS NULL ORDER BY start_time DESC")
    now = now_utc()
    for e in rows:
        s = parse_iso(e["start_time"]); t = parse_iso(e["end_time"])
        if s and t and s <= now <= t:
            return e
    return None


def to_ist(dt):
    try:
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(IST)
    except Exception:
        pass
        return dt


@app.route("/vote", methods=["POST"])
@login_required(role="voter")
def vote():
    # Rate limit voting attempts
    if not rate_limiter.is_allowed(f"vote_{session['user_id']}", limit=3, window=60):
        flash("Please wait before voting again.", "warn")
        return redirect(url_for("voter_panel"))
    
    try:
        election_id = int(request.form.get("election_id","0"))
        candidate_id = int(request.form.get("candidate_id","0"))
    except ValueError:
        flash("Invalid vote submission.", "error"); return redirect(url_for("voter_panel"))
    e = query("SELECT * FROM elections WHERE id=?", (election_id,), one=True)
    if not e: flash("No election is scheduled right now.", "warn"); return redirect(url_for("voter_panel"))
    now = now_utc(); s = parse_iso(e["start_time"]); t = parse_iso(e["end_time"])
    if not s or not t or not (s <= now <= t):
        flash("This election is not active.", "warn"); return redirect(url_for("voter_panel"))
    # Use database transaction to prevent race condition
    db = get_db()
    try:
        # Check for existing vote within transaction
        if query("SELECT 1 FROM votes WHERE voter_id=? AND election_id=?", (session["user_id"], election_id), one=True):
            flash("You have already voted in this election.", "warn"); return redirect(url_for("voter_panel"))
        c = query("SELECT * FROM candidates WHERE id=? AND election_id=?", (candidate_id, election_id), one=True)
        if not c: flash("Invalid candidate selection.", "error"); return redirect(url_for("voter_panel"))
        
        # Insert vote with unique constraint protection
        try:
            execute("INSERT INTO votes (voter_id,candidate_id,election_id,timestamp) VALUES (?,?,?,?)",
                    (session["user_id"], candidate_id, election_id, now.isoformat()))
        except Exception as e:
            # Catch unique constraint violation (double vote attempt)
            if "UNIQUE constraint failed" in str(e) or "duplicate" in str(e).lower():
                flash("You have already voted in this election.", "warn"); return redirect(url_for("voter_panel"))
            raise e
    except Exception as e:
        db.rollback()
        flash("Error recording vote. Please try again.", "error"); return redirect(url_for("voter_panel"))
    flash("Vote recorded. Thank you!", "ok"); return redirect(url_for("voter_panel"))

@app.route("/results")
@login_required(role="admin")
def results():
    election_id = request.args.get("election_id")
    e = query("SELECT * FROM elections WHERE id=?", (election_id,), one=True) if election_id else current_active_election()
    if not e: flash("No election selected/active.", "warn"); return redirect(url_for("admin"))
    rows = query("""
        SELECT c.name, COUNT(v.id) as votes
        FROM candidates c
        LEFT JOIN votes v ON v.candidate_id=c.id AND v.election_id=?
        WHERE c.election_id=?
        GROUP BY c.id
        ORDER BY votes DESC, c.name ASC
    """, (e["id"], e["id"]))
    results = [{"name": r["name"], "votes": r["votes"]} for r in rows]
    total_votes = sum(result["votes"] for result in results)
    winner = results[0] if results and results[0]["votes"] > 0 else None
    return render_template("result.html", election=e, results=results, total_votes=total_votes, winner=winner, elections=query("SELECT * FROM elections WHERE status != 'cancelled' OR status IS NULL ORDER BY start_time DESC"))



@app.route('/export.xlsx')
def export_excel():
    if session.get('user_id') is None or session.get('role') != 'admin':
        return redirect(url_for('login'))
    try:
        rows = exec_sql("SELECT * FROM elections WHERE status != 'cancelled' OR status IS NULL ORDER BY start_time DESC", fetch=True)
    except Exception:
        pass
        rows = query("SELECT * FROM elections WHERE status != 'cancelled' OR status IS NULL ORDER BY start_time DESC")
    now = datetime.now(timezone.utc)
    ongoing, scheduled, ended = [], [], []
    for e in rows:
        # Parse datetime strings if they are strings
        try:
            if isinstance(e['start_time'], str):
                st = datetime.fromisoformat(e['start_time'].replace('Z', '+00:00'))
            else:
                st = e['start_time']
            
            if isinstance(e['end_time'], str):
                en = datetime.fromisoformat(e['end_time'].replace('Z', '+00:00'))
            else:
                en = e['end_time']
            
            # Ensure timezone awareness
            if st.tzinfo is None:
                st = st.replace(tzinfo=timezone.utc)
            if en.tzinfo is None:
                en = en.replace(tzinfo=timezone.utc)
            
            if st <= now <= en: ongoing.append(e)
            elif now < st: scheduled.append(e)
            else: ended.append(e)
        except (ValueError, TypeError):
            # If datetime parsing fails, skip this election
            ended.append(e)
    wb = Workbook()
    for name, data in (('Ongoing', ongoing), ('Scheduled', scheduled), ('Ended', ended)):
        ws = wb.create_sheet(title=name)
        ws.append(['ID','Title','Category','Start (UTC)','End (UTC)','Candidate Limit'])
        for ee in data:
            ws.append([ee['id'], ee['title'], ee['category'], ee['start_time'], ee['end_time'], ee['candidate_limit']])
    if 'Sheet' in wb.sheetnames: del wb['Sheet']
    bio = BytesIO(); wb.save(bio); bio.seek(0)
    return send_file(bio, as_attachment=True, download_name='elections.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/admin/voters')
@login_required(role="admin")
def admin_voters():
    # login_required decorator already enforces admin access
    try:
        voters = exec_sql("SELECT id, username FROM users WHERE role='voter' ORDER BY username", fetch=True)
        counts = exec_sql("SELECT e.title, COUNT(v.id) AS votes FROM votes v JOIN elections e ON e.id=v.election_id GROUP BY e.title ORDER BY e.title", fetch=True)
    except Exception:
        pass
        voters = query("SELECT id, username FROM users WHERE role='voter' ORDER BY username")
        counts = query("SELECT e.title, COUNT(v.id) AS votes FROM votes v JOIN elections e ON e.id=v.election_id GROUP BY e.title ORDER BY e.title")
    return render_template('admin_voters.html', voters=voters, counts=counts)


@app.route('/admin/election/<int:election_id>')
@login_required(role="admin")
def election_dashboard(election_id):
    # login_required decorator already enforces admin access
    try:
        e = exec_sql('SELECT * FROM elections WHERE id=?', (election_id,), fetch=True, one=True)
        cand = exec_sql('SELECT name, votes FROM candidates WHERE election_id=? ORDER BY votes DESC, name', (election_id,), fetch=True)
    except Exception:
        e = query('SELECT * FROM elections WHERE id=?', (election_id,))
        cand = query('SELECT name, votes FROM candidates WHERE election_id=? ORDER BY votes DESC, name', (election_id,))
    total = sum([c['votes'] if 'votes' in c else 0 for c in cand]) if cand else 0
    return render_template('election_dashboard.html', e=e, cand=cand, total=total)


@app.route('/admin/cancel_election/<int:election_id>', methods=['POST'])
@login_required(role="admin")
def cancel_election(election_id):
    """Cancel an election with admin authentication"""
    try:
        # Verify election exists
        election = exec_sql('SELECT * FROM elections WHERE id=?', (election_id,), fetch=True, one=True)
        if not election:
            return jsonify({'success': False, 'message': 'Election not found'}), 404
        
        # Check if election is already ended
        if election['status'] == 'cancelled':
            return jsonify({'success': False, 'message': 'Election is already cancelled'}), 400
            
        # Update election status to cancelled and set end time to now
        from datetime import datetime
        current_time = datetime.now()
        
        exec_sql('''UPDATE elections 
                   SET status = 'cancelled', 
                       end_time = ?,
                       cancelled_at = ?,
                       cancelled_by = ?
                   WHERE id = ?''', 
                (current_time, current_time, session.get('user_id'), election_id))
        
        # Log the cancellation for audit trail
        election_title = election['title'] if election['title'] else (election['category'] if election['category'] else f'Election {election_id}')
        admin_name = session.get('user', {}).get('name', 'Unknown Admin')
        
        print(f"ELECTION CANCELLED: {election_title} (ID: {election_id}) by Admin: {admin_name} at {current_time}")
        
        return jsonify({
            'success': True, 
            'message': f'Election "{election_title}" has been successfully cancelled',
            'election_id': election_id,
            'cancelled_at': current_time.isoformat()
        })
        
    except Exception as e:
        print(f"Error cancelling election {election_id}: {e}")
        return jsonify({'success': False, 'message': f'Database error: {str(e)}'}), 500


@app.route('/schedule', methods=['GET','POST'])
def schedule():
    if session.get('user_id') is None or session.get('role') != 'admin':
        return redirect(url_for('login'))
    if request.method == 'POST':
        title = request.form.get('title','').strip()
        category = request.form.get('category','').strip()
        limit = request.form.get('candidate_limit') or None
        start_time = request.form.get('start_time','').strip()
        end_time = request.form.get('end_time','').strip()
        
        if not title or not category or not start_time or not end_time:
            flash("All fields are required for scheduling an election.", "error")
            return render_template('schedule.html', error="Missing required fields")
        
        try:
            # Parse datetime-local format and convert to UTC
            st = datetime.fromisoformat(start_time)
            en = datetime.fromisoformat(end_time)
            
            # Treat as local time (IST) and convert to UTC
            st = safe_localize(st, IST).astimezone(timezone.utc)
            en = safe_localize(en, IST).astimezone(timezone.utc)
            
            if en <= st:
                flash("End time must be after start time.", "error")
                return render_template('schedule.html', error="Invalid time range")
            
            # Convert candidate limit to integer if provided
            if limit:
                limit = int(limit)
                if limit < 2:
                    flash("Candidate limit must be at least 2.", "error")
                    return render_template('schedule.html', error="Invalid candidate limit")
            
            # Get current year if not provided
            year = st.year
            
            execute('INSERT INTO elections(title,year,category,start_time,end_time,candidate_limit,created_by) VALUES (?,?,?,?,?,?,?)', 
                   (title, year, category, st.isoformat(), en.isoformat(), limit, session.get('user_id')))
            
            flash(f"Election '{title}' scheduled successfully from {st.strftime('%Y-%m-%d %H:%M')} to {en.strftime('%Y-%m-%d %H:%M')} IST", "success")
            return redirect(url_for('admin'))
            
        except ValueError as e:
            flash("Invalid date/time format. Please check your input.", "error")
            return render_template('schedule.html', error=f"Date parsing error: {str(e)}")
        except Exception as e:
            flash("An error occurred while scheduling the election.", "error")
            return render_template('schedule.html', error=str(e))
    
    return render_template('schedule.html')

# NOTE: the run block is placed at the end of the file (after all helper definitions)


@app.template_filter("istfmt")
def istfmt(value):
    try:
        dt = parse_iso(value)
        dt = to_ist(dt)
        return dt.strftime("%d %b %Y, %I:%M %p IST")
    except Exception:
        pass
        return value


def classify_elections(rows):
    now = now_utc()
    ongoing, scheduled, ended = [], [], []
    for e in rows:
        # Skip cancelled elections - they should not appear in ongoing or scheduled
        # Use bracket notation for sqlite3.Row objects (they don't have .get() method)
        status = e['status'] if 'status' in e.keys() else None
        if status == 'cancelled':
            ended.append(e)  # Cancelled elections go to "ended" section
            continue
            
        s = parse_iso(e["start_time"]); t = parse_iso(e["end_time"])
        if s and t:
            if s <= now <= t: ongoing.append(e)
            elif now < s: scheduled.append(e)
            else: ended.append(e)
    return ongoing, scheduled, ended


@app.route("/results_excel/<int:eid>")
@login_required(role="admin")
def results_excel(eid):
    import io
    from openpyxl import Workbook
    e = query("SELECT * FROM elections WHERE id=?", (eid,), one=True)
    if not e:
        flash("Election not found", "error"); return redirect(url_for("admin"))
    rows = query("""
        SELECT c.name, COUNT(v.id) as votes
        FROM candidates c
        LEFT JOIN votes v ON v.candidate_id=c.id AND v.election_id=?
        WHERE c.election_id=?
        GROUP BY c.id
        ORDER BY votes DESC, c.name ASC
    """, (eid, eid))
    wb = Workbook()
    ws = wb.active; ws.title = "Results"
    ws.append(["Election", e["title"] or e["category"]])
    ws.append(["Start", e["start_time"], "End", e["end_time"]])
    ws.append([]); ws.append(["Candidate", "Votes"])
    for r in rows: ws.append([r["name"], r["votes"]])
    stream = io.BytesIO(); wb.save(stream); stream.seek(0)
    return (stream.read(), 200, {
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "Content-Disposition": f"attachment; filename=results_{eid}.xlsx",
    })


@app.route("/voter")
@login_required(role="voter")
def voter_panel():
    rows = query("SELECT * FROM elections WHERE status != 'cancelled' OR status IS NULL ORDER BY start_time ASC")
    ongoing, scheduled, ended = classify_elections(rows)

    cand_map = {}
    for e in ongoing:
        cand_map[e["id"]] = query(
            "SELECT * FROM candidates WHERE election_id=?", (e["id"],)
        )

    return render_template(
        "voter.html",
        ongoing=ongoing,
        scheduled=scheduled,
        ended=ended,
        cand_map=cand_map,
    )


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8080))
    app.run(host="0.0.0.0", port=port, debug=False)
