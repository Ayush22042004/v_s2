from datetime import datetime, timezone

def parse_utc_iso(dt_str):
    if not dt_str:
        return None
    # Accept ...Z or +00:00
    return datetime.fromisoformat(dt_str.replace('Z','+00:00'))

from io import BytesIO
from openpyxl import Workbook
from flask import send_file, redirect, url_for, session, render_template
from datetime import datetime, timezone

@app.route("/results")
def results():
    u = session.get("user")
    if not u or u.get("role") != "admin":
        return redirect(url_for("login"))
    rows = query("""
        SELECT e.title AS election_title, c.name AS candidate_name, c.votes AS votes
        FROM candidates c
        JOIN elections e ON e.id = c.election_id
        WHERE e.end_time < CURRENT_TIMESTAMP
        ORDER BY e.end_time DESC, votes DESC, candidate_name
    """)  # rename 'query' to your helper if needed
    return render_template("results.html", rows=rows)

@app.route("/export.xlsx")
def export_excel():
    u = session.get("user")
    if not u or u.get("role") != "admin":
        return redirect(url_for("login"))
    rows = query("SELECT * FROM elections ORDER BY start_time DESC")  # rename if needed
    now = datetime.now(timezone.utc)
    ongoing, scheduled, ended = [], [], []
    for e in rows:
        st, en = e["start_time"], e["end_time"]
        if st <= now <= en: ongoing.append(e)
        elif now < st: scheduled.append(e)
        else: ended.append(e)
    wb = Workbook()
    for name, data in (("Ongoing", ongoing), ("Scheduled", scheduled), ("Ended", ended)):
        ws = wb.create_sheet(title=name)
        ws.append(["ID","Title","Category","Start (UTC)","End (UTC)","Candidate Limit"])
        for e in data:
            ws.append([e["id"], e["title"], e["category"], e["start_time"], e["end_time"], e.get("candidate_limit")])
    if "Sheet" in wb.sheetnames: del wb["Sheet"]
    bio = BytesIO(); wb.save(bio); bio.seek(0)
    return send_file(bio, as_attachment=True, download_name="elections.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
